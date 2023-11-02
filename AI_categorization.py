import sys

print(sys.executable)
print(sys.version)

import json
import pandas as pd
import os
import pprint
import openpyxl
from openpyxl.utils import get_column_letter
import openai
from dotenv import load_dotenv

# Make sure to load the environment variables from a .env file or the environment
load_dotenv()

# Retrieve the API key from an environment variable
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")


SOURCE_DIR = "ConsolidatedReports"
OUTPUT_PATH_CSV = "ConsolidatedReports/AI_Combined.csv"
OUTPUT_PATH_XLSX = "ConsolidatedReports/AI_Combined.xlsx"

# Sample constants for categories and the prompt
CATEGORIES = [
    "Office Supplies",
    "Internet Expenses",
    "Equipment Maintenance",
    "Automobile",
    "Outside Services",
    "Parking and Tolls",
    "Computer Expenses",
    "Business Travel Expenses",
    "Client Gifts",
    "Advertising",
    "Personal",
    "Computer Equipment",
    "Telecom",
    "Office Rent",
    "Utilities",
    "Office Furniture",
    "Electronics",
    "Marketing and Promotion",
    "Professional Fees (Legal, Accounting)",
    "Software Licenses and Subscriptions",
    "Employee Benefits and Perks",
    "Meals and Entertainment",
    "Shipping and Postage",
]

PROMPT_TEMPLATE = """Based on the description of the following item, 
please categorize it into one of the predefined tax write-off categories: {categories}.
The item description is: {{item_description}}"""


# Initialize a placeholder DataFrame for demonstration
df_amazon = pd.DataFrame(
    {
        "TransactionID": [1, 2, 3],
        "Item_Description": ["Laptop", "Office Chair", "Printer Ink"],
        "Amount": [1000, 200, 50],
    }
)


# Define a function to read a CSV file into a pandas DataFrame
def read_dataset(file_path):
    """
    Reads a CSV file from a given file path into a pandas DataFrame.

    Parameters:
    file_path (str): The path to the CSV file.

    Returns:
    pd.DataFrame: The DataFrame containing the data from the CSV file.
    """
    try:
        df = pd.read_csv(file_path)
        return df
    except FileNotFoundError:
        print("File not found. Please provide a valid file path.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# Simulating a function to apply initial categorizations using the OpenAI API
def categorize_transaction(df, column_name):
    """
    Applies initial categorizations to transactions based on item descriptions.

    Parameters:
    df (pd.DataFrame): The DataFrame containing the transactions.
    column_name (str): The name of the column containing item descriptions.

    Returns:
    pd.DataFrame: The DataFrame with an additional column for categories.
    """
    # Placeholder for categories, this would be replaced by OpenAI API output in a real application
    categories = ["Electronics", "Office Furniture", "Office Supplies"]

    # Adding the categories to the DataFrame
    df["Category"] = categories

    return df


# Define a function to export a DataFrame to an Excel sheet
def export_to_excel(df, file_name, sheet_name):
    """
    Exports a pandas DataFrame to a new sheet in an existing or new Excel workbook.

    Parameters:
    df (pd.DataFrame): The DataFrame to export.
    file_name (str): The name of the Excel file.
    sheet_name (str): The name of the Excel sheet.

    Returns:
    None: Saves the DataFrame to an Excel sheet.
    """
    try:
        # Check if the Excel file already exists
        try:
            book = openpyxl.load_workbook(file_name)
            writer = pd.ExcelWriter(file_name, engine="openpyxl", mode="a")
            writer.book = book
        except FileNotFoundError:
            writer = pd.ExcelWriter(file_name, engine="openpyxl", mode="w")

        # Export the DataFrame to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()

        print(f"Data successfully saved to {sheet_name} in {file_name}.")
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# For demonstration, let's assume we are saving to a file named 'Tax_Write_Offs.xlsx' and a sheet named 'Amazon_Data'
# export_to_excel(df_amazon_categorized, 'Tax_Write_Offs.xlsx', 'Amazon_Data')  # Uncomment this line when you actually want to save the file

# The function is ready but we won't actually run it here as file operations are for demonstration purposes only


# Applying initial categorizations to the placeholder DataFrame
# df_amazon_categorized = categorize_transaction(df_amazon, "Item_Description")
# df_amazon_categorized


# Define a function to simulate AI categorization using OpenAI API
def simulate_openai_categorization(description, prompt):
    """
    Simulates the categorization of a transaction description using OpenAI API.

    Parameters:
    description (str): The transaction description to categorize.
    prompt (str): The OpenAI prompt for categorization.

    Returns:
    str: The category as determined by the simulated OpenAI API.
    """
    # Simulating API call by selecting a category based on keywords (this is a placeholder)
    keywords_to_categories = {
        "Laptop": "Electronics",
        "Office Chair": "Office Furniture",
        "Printer Ink": "Office Supplies",
    }

    # For demonstration, let's assume we are adding dropdowns to 'Final_Category' column in 'Tax_Write_Offs.xlsx' and sheet 'Amazon_Data'
    # The column for 'Final_Category' would be 'E' based on our placeholder DataFrame
    # add_dropdowns_to_excel('Tax_Write_Offs.xlsx', 'Amazon_Data', 'E', categories_for_dropdown)  # Uncomment this line when you actually want to add dropdowns

    # The function is ready but we won't actually run it here as file operations are for demonstration purposes only
    # Simulating the creation of a main program script to execute the entire workflow
    return keywords_to_categories.get(description, "Uncategorized")


def simulate_openai_categorization_advanced(description, prompt):
    # This function simulates the categorization process of the OpenAI API with custom instructions
    # Replace this with an actual API call in your implementation
    # The prompt variable would be passed to the API call
    category_mapping = {
        "Laptop": "Computer Equipment",
        "Internet Bill": "Internet Expenses",
        "Office Chair": "Office Furniture",
    }
    return category_mapping.get(description, "Unknown")


def categorize_with_openai(description, prompt):
    # Replace 'your_openai_api_key' with your actual OpenAI API key
    openai.api_key = "your_openai_api_key"

    # This is where you would construct your actual prompt using the description and any custom instructions
    full_prompt = f"{prompt}\nThe item description is: {description}"

    # Make the OpenAI API call with the prompt
    response = openai.Completion.create(
        engine="davinci",
        prompt=full_prompt,
        max_tokens=60,
        n=1,
        stop=None,
        temperature=0.3,
    )

    # Extract the category from the response
    category = response.choices[0].text.strip()

    # Return the category
    return category


# Define a function to integrate AI categorizations into the data pipeline
def integrate_ai_categorization(df, ai_category_column):
    """
    Integrates the AI-driven categorizations into the DataFrame.

    Parameters:
    df (pd.DataFrame): The DataFrame containing transactions and initial categories.
    ai_category_column (str): The name of the column containing AI-driven categories.

    Returns:
    pd.DataFrame: The DataFrame with AI-driven categorizations integrated.
    """
    # For demonstration, we'll simply copy the AI categories to the 'Final_Category' column
    # In a real-world application, you might want to apply some logic to resolve discrepancies between initial and AI categorizations
    df["Final_Category"] = df[ai_category_column]

    return df


# Integrating AI-driven categorizations into the placeholder DataFrame
# df_amazon_final = integrate_ai_categorization(df_amazon_categorized, "AI_Category")
# df_amazon_final


# Importing necessary modules for adding dropdowns in Excel
from openpyxl.worksheet.datavalidation import DataValidation


# Define a function to add dropdown lists to an Excel sheet
def add_dropdowns_to_excel(file_name, sheet_name, column_letter, categories):
    """
    Adds dropdown lists to a specific column in an Excel sheet.

    Parameters:
    file_name (str): The name of the Excel file.
    sheet_name (str): The name of the Excel sheet.
    column_letter (str): The letter representing the column to which to add dropdowns.
    categories (list): The list of categories to include in the dropdowns.

    Returns:
    None: Adds dropdowns to the Excel sheet and saves the changes.
    """
    try:
        # Load the Excel workbook and sheet
        book = openpyxl.load_workbook(file_name)
        sheet = book[sheet_name]

        # Create a data validation object for the dropdowns
        dv = DataValidation(
            type="list", formula1=f'"{",".join(categories)}"', showDropDown=True
        )

        # Apply the data validation to the column
        for row in range(
            2, sheet.max_row + 1
        ):  # Starting from row 2 to skip the header
            sheet[f"{column_letter}{row}"].validation = dv

        # Save the changes
        book.save(file_name)

        print(
            f"Dropdowns successfully added to column {column_letter} in {sheet_name} of {file_name}."
        )
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# Simulating a function to scan a directory for Excel or CSV files
def scan_directory_for_files(directory_path, file_types=[".csv"]):
    """
    Scans a directory and returns a list of Excel or CSV files.

    Parameters:
    directory_path (str): The path to the directory to scan.
    file_types (list): A list of file extensions to look for (default is ['.csv', '.xlsx']).

    Returns:
    list: A list of file paths that match the specified file types.
    """
    # For demonstration, we'll return a hard-coded list as we can't access the file system here
    # return ["BofA.csv", "Chase.csv", "Amazon.csv"]

    # In a real application, you would use the following code:
    return [
        os.path.join(directory_path, f)
        for f in os.listdir(directory_path)
        if os.path.splitext(f)[1] in file_types
    ]


# Update the read_and_standardize_file function to handle the actual headers involved in each dataset


def read_and_standardize_file(file_path):
    """
    Reads and standardizes a file based on its type, taking into account the actual headers involved in each dataset.

    Parameters:
    file_path (str): The path to the file.

    Returns:
    pd.DataFrame: A standardized DataFrame containing the data from the file.
    """
    try:
        # Determine the file type based on its extension
        file_type = os.path.splitext(file_path)[1]

        # Read the file into a DataFrame
        if file_type == ".csv":
            df = pd.read_csv(file_path)
        elif file_type == ".xlsx":
            df = pd.read_excel(file_path)
        else:
            print(f"Unsupported file type: {file_type}")
            return None

        # Identify the source (e.g., BofA, Chase, Amazon) based on the file name
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        source = None
        if "BofA" in file_name:
            source = "BofA"
        elif "Chase" in file_name:
            source = "Chase"
        elif "Amazon" in file_name:
            source = "Amazon"

        if source is None:
            print(f"Unknown source: {file_name}")
            return None

        df["Source"] = source

        # Standardize the column names based on the actual headers for each source
        if source == "BofA":
            df.rename(
                columns={
                    "Transaction Date": "Transaction_Date",
                    "Posting Date": "Posting_Date",
                    "Description": "Item_Description",
                    "Reference Number": "Reference_Number",
                    "Account Number": "Account_Number",
                    "Amount": "Amount",
                    "Statement Date": "Statement_Date",
                },
                inplace=True,
            )
            df["TransactionID"] = df.index + 1  # Generate a TransactionID

        elif source == "Chase":
            df.rename(
                columns={
                    "Date of Transaction": "Transaction_Date",
                    "Merchant Name or Transaction Description": "Item_Description",
                    "$ Amount": "Amount",
                    "Statement Date": "Statement_Date",
                    "Date": "Posting_Date",
                },
                inplace=True,
            )
            df["TransactionID"] = df.index + 1  # Generate a TransactionID

        elif source == "Amazon":
            # Assume the 'Items' column contains the list of objects
            # Each object has a 'description' and 'price' (this is a simplification for demonstration purposes)
            sub_orders = []
            for i, row in df.iterrows():
                items = eval(
                    row["Items"]
                )  # Assuming the list of objects is stored as a string representation
                for item in items:
                    sub_order = row.copy()
                    sub_order["Item_Description"] = item["Description"]
                    sub_order["Amount"] = item["Price"]
                    sub_order[
                        "TransactionID"
                    ] = f"{row['Order Number']}_{item['Description']}"
                    sub_orders.append(sub_order)

            df = pd.DataFrame(sub_orders)

            # Check if a Gift Card was used
            for i, row in df.iterrows():
                if row["Amount"] == 0.0 and row["Gift_Card_Amount"] > 0:
                    df.at[i, "Amount"] = row["Gift_Card_Amount"]

        return df

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# The function is now updated to handle the actual headers and generate a TransactionID if necessary
# Uncomment the next line when you want to run the function with an actual file
# read_and_standardize_file_actual_headers('BofA_Sample.csv')


# Simulating a function to export each DataFrame to its own Excel sheet
def export_multiple_dfs_to_excel(dfs, file_name):
    """
    Exports multiple pandas DataFrames to their own sheets in an Excel workbook.

    Parameters:
    dfs (dict): A dictionary containing the DataFrames to export. The keys will be used as sheet names.
    file_name (str): The name of the Excel file to which to export the DataFrames.

    Returns:
    None: Saves each DataFrame to its own sheet in the Excel workbook.
    """
    try:
        # Initialize an Excel writer object
        writer = pd.ExcelWriter(file_name, engine="openpyxl")

        # Loop through the dictionary and save each DataFrame to its own sheet
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Save the Excel file
        writer.save()

        print(f"DataFrames successfully saved to their own sheets in {file_name}.")
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# For demonstration, let's assume we are saving to a file named 'Consolidated_Reports.xlsx'
# export_multiple_dfs_to_excel(standardized_dfs, 'Consolidated_Reports.xlsx')  # Uncomment this line when you actually want to save the file

# The function is ready but we won't actually run it here as file operations are for demonstration purposes only


# Simulating a function to add dropdown lists to multiple sheets in an Excel workbook
def add_dropdowns_to_multiple_sheets(file_name, sheet_column_mapping, categories):
    """
    Adds dropdown lists to specified columns in multiple sheets of an Excel workbook.

    Parameters:
    file_name (str): The name of the Excel file.
    sheet_column_mapping (dict): A dictionary mapping sheet names to the columns where dropdowns should be added.
    categories (list): The list of categories to include in the dropdowns.

    Returns:
    None: Adds dropdowns to the specified columns in each sheet and saves the changes.
    """
    try:
        # Load the Excel workbook
        book = openpyxl.load_workbook(file_name)

        # Create a data validation object for the dropdowns
        dv = DataValidation(
            type="list", formula1=f'"{",".join(categories)}"', showDropDown=True
        )

        # Loop through the sheets and add dropdowns
        for sheet_name, column_letter in sheet_column_mapping.items():
            sheet = book[sheet_name]
            for row in range(
                2, sheet.max_row + 1
            ):  # Starting from row 2 to skip the header
                sheet[f"{column_letter}{row}"].validation = dv

        # Save the changes
        book.save(file_name)

        print(f"Dropdowns successfully added to multiple sheets in {file_name}.")
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


# For demonstration, let's assume we are adding dropdowns to 'Final_Category' column in 'Consolidated_Reports.xlsx'
# The column for 'Final_Category' would be 'D' based on our placeholder DataFrames
# sheet_column_mapping = {sheet_name: 'D' for sheet_name in standardized_dfs.keys()}
# add_dropdowns_to_multiple_sheets('Consolidated_Reports.xlsx', sheet_column_mapping, categories_for_dropdown)  # Uncomment this line when you actually want to add dropdowns


def batch_categorize_items(descriptions, categories, api_key):
    openai.api_key = api_key
    # Construct the JSON input for the prompt
    json_input = {"descriptions": descriptions, "categories": categories}
    json_input_str = json.dumps(
        json_input, indent=2
    )  # Make it human-readable for the AI

    # Create a structured prompt with JSON input and expected JSON output
    prompt = (
        "Return a JSON object that categorizes the following descriptions of various transactions. "
        "Please categorize each transaction by returning a JSON object with each "
        "description mapped to the most suitable category from the category in the list. Do not include any other \n"
        "content in your resopnse because your response will be assumed to be pure JSON object for python processing. \n"
        "Here is the the JSON you should use as your input json object with descriptions and input json object with categories:"
        f"Input:\n{json_input_str}\n"
    )

    # Make the OpenAI API call
    # Construct the payload
    # payload = {
    #     "engine": "davinci",
    #     "prompt": prompt,
    #     "max_tokens": 1024,  # Adjust based on needs
    #     "n": 1,
    #     "stop": None,
    #     "temperature": 0,
    # }

    # # Print the payload for debugging
    # print("Payload for OpenAI API:")
    # print(json.dumps(payload, indent=2))

    # # Make the API call
    # response = openai.Completion.create(
    #     **payload, api_key=api_key  # Using dictionary unpacking to pass the parameters
    # )
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt},
        ],
    )
    # print(response)
    # Extract and print the relevant data
    answer = response["choices"][0]["message"]["content"]

    # Attempt to parse the response as JSON
    try:
        categorized_results = json.loads(answer)
    except json.JSONDecodeError:
        categorized_results = {"error": "Invalid JSON response"}

    return categorized_results


# Basic OPEN AI Call for testing
def get_gpt_response(api_key):
    openai.api_key = api_key
    # Craft the prompt
    list1 = [1, 2, 3, 4, 5]
    list2 = [6, 7, 8, 9, 10, 11]
    prompt = f"Calculate the length of the following two lists: {list1} and {list2}."

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt},
        ],
    )
    # print(response)
    # Extract and print the relevant data
    answer = response["choices"][0]["message"]["content"]
    print(answer)
    return answer


def main():
    """
    Main function to execute the entire data processing workflow.
    """
    try:
        print("Starting the data processing workflow...")

        # Step 1: Scan the directory for files, need to exclude AI files
        print("1: Scanning the directory for CSV files...")
        file_list = scan_directory_for_files(SOURCE_DIR)
        print(f"Found files: {file_list}")

        # Step 2: Read and standardize each file
        print("2: Reading and standardizing files...")
        # Initialize an empty DataFrame to hold the consolidated data
        consolidated_df = pd.DataFrame()
        standardized_dfs = {}
        for file in file_list:
            # standardized_dfs[file] = read_and_standardize_file(file)
            # pprint.pprint(standardized_dfs[file])
            standardized_df = read_and_standardize_file(file)
            standardized_df["Source_File"] = file
            # pprint.pprint(standardized_df)
            consolidated_df = pd.concat(
                [consolidated_df, standardized_df], ignore_index=True
            )
        print("Standardization complete.")
    except Exception as e:
        print(f"An error occurred in steps 1-2: {e}")

    try:
        # print("OPEN AI TEST:")
        # result = get_gpt_response(api_key=OPENAI_API_KEY)
        # pprint.pprint(result)
        # print("END TEST")

        # Load your consolidated DataFrame (for testing, this will be a small subset)
        # In practice, you would load this from your CSV/XLSX or database
        consolidated_df = pd.DataFrame(
            {
                "Item_Description": ["Laptop", "Internet Bill", "Office Chair", "UBER"],
                "Amount": [1200.00, 60.00, 250.00, 25],
            }
        )

        # Your predefined categories
        categories = [
            "Office Supplies",
            "Internet Expenses",
            "Equipment Maintenance",
            "Transportation",
            "Computer",
        ]

        # Perform batch categorization using the OpenAI API
        categorized_results = batch_categorize_items(
            descriptions=consolidated_df["Item_Description"].tolist(),
            categories=categories,
            api_key=OPENAI_API_KEY,
        )

        print("///RESULTS///")
        pprint.pprint(categorized_results)
        print("///END RESULTS///")

        # Handle the results
        if "error" in categorized_results:
            print("An error occurred with the OpenAI API call:")
            pprint.pprint(categorized_results)
        else:
            # Update the DataFrame with the categorized results
            for description, category in categorized_results.items():
                consolidated_df.loc[
                    consolidated_df["Item_Description"] == description, "AI_Category"
                ] = category

        print("Categorization completed. Here are the results:")
        pprint.pprint(consolidated_df)

    except Exception as e:
        print(f"An error occurred in step 3: {e}")


# Run the main function
if __name__ == "__main__":
    main()
