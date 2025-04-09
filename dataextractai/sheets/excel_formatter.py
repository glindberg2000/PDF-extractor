"""Excel formatter for creating rich transaction reports with validation and charts."""

import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import json
import os
import logging
import sqlite3
from ..db.client_db import ClientDB
from ..utils.constants import CATEGORY_ID_TO_NAME, ALLOWED_WORKSHEETS
from ..utils.tax_categories import (
    TAX_WORKSHEET_CATEGORIES,
    EXCLUDED_FROM_6A,
    should_exclude_from_worksheet,
)

logger = logging.getLogger(__name__)


class ExcelReportFormatter:
    """Creates rich Excel reports with validation, charts, and summaries."""

    def __init__(self):
        """Initialize the Excel formatter with styles."""
        self.header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        self.header_font = Font(color="FFFFFF", bold=True)
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Initialize category mappings
        self.category_id_to_name = {}
        self.tax_category_id_to_name = {}

        # Load category mappings from database
        self._load_category_mappings()

        # Define columns to show/hide
        self.visible_columns = [
            # Transaction Info
            "transaction_id",
            "transaction_date",
            "description",
            "amount",
            "normalized_amount",
            "source",
            # Payee Info (Pass 1)
            "payee",
            "payee_confidence",
            "payee_reasoning",
            # Category Info (Pass 2)
            "category",
            "category_confidence",
            "category_reasoning",
            "expense_type",
            "business_percentage",
            "business_context",
            # Tax Info (Pass 3)
            "classification",
            "classification_confidence",
            "classification_reasoning",
            "tax_category",
            "tax_subcategory",
            "worksheet",
            "tax_worksheet_line_number",
            "tax_year",
            # Review Info
            "is_reviewed",
            "review_notes",
            "last_reviewed_at",
        ]

    def _get_business_categories(self, client_name):
        """Get categories from business profile."""
        try:
            # First try to load from database
            db = ClientDB()
            profile = db.load_profile(client_name)

            if profile:
                # Combine AI-generated and custom categories
                categories = set()

                # Add main categories from hierarchy
                if "category_hierarchy" in profile:
                    categories.update(
                        profile["category_hierarchy"].get("main_categories", [])
                    )
                    # Add subcategories
                    for subcats in (
                        profile["category_hierarchy"].get("subcategories", {}).values()
                    ):
                        categories.update(subcats)

                # Add custom categories
                if "custom_categories" in profile:
                    categories.update(profile["custom_categories"])

                # Add AI-generated categories
                if "ai_generated_categories" in profile:
                    categories.update(profile["ai_generated_categories"])

                # Add standard 6A categories
                if "6A" in TAX_WORKSHEET_CATEGORIES:
                    categories.update(TAX_WORKSHEET_CATEGORIES["6A"])

                # Add "Other" category
                categories.add("Other")

                return sorted(list(categories))

            # Fallback to file-based approach if profile not in database
            profile_path = os.path.join(
                "data", "clients", client_name, "business_profile.json"
            )
            with open(profile_path, "r") as f:
                profile = json.load(f)

            # Combine AI-generated and custom categories
            categories = set()

            # Add main categories from hierarchy
            if "category_hierarchy" in profile:
                categories.update(
                    profile["category_hierarchy"].get("main_categories", [])
                )
                # Add subcategories
                for subcats in (
                    profile["category_hierarchy"].get("subcategories", {}).values()
                ):
                    categories.update(subcats)

            # Add custom categories
            if "custom_categories" in profile:
                categories.update(profile["custom_categories"])

            # Add AI-generated categories
            if "ai_generated_categories" in profile:
                categories.update(profile["ai_generated_categories"])

            # Add standard 6A categories
            if "6A" in TAX_WORKSHEET_CATEGORIES:
                categories.update(TAX_WORKSHEET_CATEGORIES["6A"])

            # Add "Other" category
            categories.add("Other")

            return sorted(list(categories))
        except Exception as e:
            logger.warning(f"Error reading business profile: {e}")
            # Fallback to standard 6A categories plus basic defaults
            categories = [
                "Income",
                "Business Expense",
                "Personal Expense",
                "Transfer",
                "Other",
            ]
            if "6A" in TAX_WORKSHEET_CATEGORIES:
                categories.extend(TAX_WORKSHEET_CATEGORIES["6A"])
            return sorted(list(set(categories)))

    def _normalize_tax_category(self, category: str) -> str:
        """Normalize tax category names to prevent duplicates."""
        if not category:
            return "Other expenses"

        # Convert to lowercase for comparison
        category = category.lower().strip()

        # Remove common variations
        category = category.replace("expense", "").replace("expenses", "").strip()

        # Standard mappings
        mappings = {
            "advertising": "Advertising",
            "car and truck": "Car and truck expenses",
            "commissions and fee": "Commissions and fees",
            "contract labor": "Contract labor",
            "employee benefit": "Employee benefit programs",
            "insurance": "Insurance (other than health)",
            "interest": "Interest",
            "legal and professional": "Legal and professional services",
            "office": "Office expenses",
            "rent or lease": "Rent or lease",
            "repairs and maintenance": "Repairs and maintenance",
            "supplies": "Supplies",
            "taxes and license": "Taxes and licenses",
            "travel and meal": "Travel and meals",
            "utilities": "Utilities",
            "wages": "Wages",
            "other": "Other expenses",
        }

        # Find the best match
        for key, value in mappings.items():
            if key in category:
                return value

        return "Other expenses"

    def _load_category_mappings(self):
        """Load category mappings from the database."""
        try:
            db = ClientDB()
            with sqlite3.connect(db.db_path) as conn:
                # Load tax categories
                cursor = conn.execute("SELECT id, name FROM tax_categories")
                for cat_id, cat_name in cursor.fetchall():
                    self.tax_category_id_to_name[cat_id] = cat_name
                    self.tax_category_id_to_name[float(cat_id)] = (
                        cat_name  # Handle float conversion
                    )
                    self.tax_category_id_to_name[str(cat_id)] = (
                        cat_name  # Handle string conversion
                    )

                # Use the shared category mapping from constants
                self.category_id_to_name = CATEGORY_ID_TO_NAME

            logger.info(
                f"Loaded {len(self.tax_category_id_to_name)} tax category mappings and {len(self.category_id_to_name)} business category mappings"
            )
        except Exception as e:
            logger.error(f"Error loading category mappings: {e}")

    def create_report(
        self,
        data: pd.DataFrame,
        output_path: str,
        client_name: str,
        use_linked_formulas: bool = False,
    ):
        """Create Excel report with validation, charts, summaries, and separate CSVs.

        Args:
            data: DataFrame with transaction data
            output_path: Path to save the Excel file
            client_name: Name of the client for profile loading
            use_linked_formulas: Whether to use linked formulas instead of raw values
        """
        # If linked formulas are requested, use the new method
        if use_linked_formulas:
            self.create_linked_excel_report(data, output_path, client_name)

            # Define the base output directory and create the CSV subdirectory
            base_output_dir = os.path.dirname(output_path)
            csv_output_dir = os.path.join(base_output_dir, "csv_sheets")

            # Still save CSV files
            self.save_csv_files(data, client_name, csv_output_dir)
            return

        # Original implementation for non-linked reports
        # Ensure the input data has the 'worksheet' column
        if "worksheet" not in data.columns:
            logger.error("Input data DataFrame is missing the 'worksheet' column.")
            raise ValueError(
                "Input data is missing the 'worksheet' column needed for report generation."
            )

        # Get categories from business profile
        categories = self._get_business_categories(client_name)
        classifications = [
            "Business",
            "Personal",
            "Unclassified",
        ]  # Default classifications

        # Fix numeric categories - convert to descriptive names if needed
        if "category" in data.columns:
            # Convert floats or ints to strings for mapping
            if data["category"].dtype in ["int64", "float64"]:
                data["category"] = data["category"].astype(str)

            # Apply mapping for numeric category values
            data["category"] = data["category"].apply(
                lambda x: (
                    self.category_id_to_name.get(x, x)
                    if x in self.category_id_to_name
                    else x
                )
            )
            logger.info("Converted numeric categories to descriptive names")

        # Fix tax category IDs - convert to actual tax category names
        if "tax_category_id" in data.columns:
            # Create a new column or update existing with descriptive names
            data["tax_category"] = data["tax_category_id"].apply(
                lambda x: self.tax_category_id_to_name.get(x, f"Unknown ({x})")
            )
            logger.info("Converted tax category IDs to category names")

        # Ensure business_percentage is 0 for Personal expenses
        if "expense_type" in data.columns and "business_percentage" in data.columns:
            mask = data["expense_type"] == "personal"
            if mask.any():
                logger.info(
                    f"Setting business_percentage to 0 for {mask.sum()} personal expenses"
                )
                data.loc[mask, "business_percentage"] = 0

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Transactions"

        # Write headers to main transactions sheet
        all_columns = list(data.columns)
        for col, header in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

            # Hide columns not in visible_columns
            if header not in self.visible_columns:
                ws.column_dimensions[get_column_letter(col)].hidden = True

        # Write data to main transactions sheet
        for row in range(len(data)):
            for col_idx, value in enumerate(data.iloc[row], 1):
                ws.cell(row=row + 2, column=col_idx, value=value)

        # Expected worksheets - these must match the literals in ClassificationResponse
        expected_worksheets = ["6A", "Vehicle", "HomeOffice", "Personal", "Unknown"]

        # Clean tax_category values if they are numeric IDs
        if "tax_category" in data.columns and data["tax_category"].dtype == "int64":
            # Try to convert numeric tax_category to string names if needed
            logger.warning("Tax categories appear to be numeric IDs instead of names")

        # Make sure expense_type is used to determine Personal items, not just worksheet
        if "expense_type" in data.columns:
            # Force worksheet to be Personal for personal expense types
            personal_mask = data["expense_type"] == "personal"
            if personal_mask.any():
                logger.info(
                    f"Setting worksheet to 'Personal' for {personal_mask.sum()} personal expense items"
                )
                data.loc[personal_mask, "worksheet"] = "Personal"

        # Ensure worksheet column contains valid values
        data["worksheet"] = data["worksheet"].fillna("Unknown").astype(str)

        # Apply exclusion logic - move excluded categories from 6A to Personal worksheet
        if "category" in data.columns and "worksheet" in data.columns:
            for idx, row in data.iterrows():
                if (
                    row["worksheet"] == "6A"
                    and isinstance(row["category"], str)
                    and should_exclude_from_worksheet(row["category"], "6A")
                ):
                    # Move to Personal worksheet
                    data.at[idx, "worksheet"] = "Personal"
                    # Set business percentage to 0
                    if "business_percentage" in data.columns:
                        data.at[idx, "business_percentage"] = 0
                    logger.info(
                        f"Excluded '{row['category']}' from 6A worksheet (txn_id: {row.get('transaction_id', idx)})"
                    )

        # Create summary worksheets for each worksheet type
        for worksheet_name in expected_worksheets:
            # Skip Unknown worksheet in summaries - it doesn't need a summary tab
            if worksheet_name == "Unknown":
                continue

            # Filter data for this worksheet
            worksheet_data = data[data["worksheet"] == worksheet_name].copy()

            if not worksheet_data.empty:
                # Create new worksheet
                ws_sheet = wb.create_sheet(worksheet_name)

                # Set up summary headers
                summary_headers = [
                    "Category",
                    "Total Amount",
                    "Count",
                    "Average Amount",
                ]
                for col, header in enumerate(summary_headers, 1):
                    cell = ws_sheet.cell(row=1, column=col, value=header)
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = self.header_alignment

                # Group by category and calculate totals
                if "category" in worksheet_data.columns:
                    # Get category summaries
                    category_summaries = (
                        worksheet_data.groupby("category")
                        .agg(
                            total_amount=("amount", "sum"),
                            count=("amount", "count"),
                        )
                        .reset_index()
                    )

                    # Add average column
                    category_summaries["average"] = (
                        category_summaries["total_amount"] / category_summaries["count"]
                    )

                    # Sort by total amount (descending)
                    category_summaries = category_summaries.sort_values(
                        "total_amount", ascending=False
                    )

                    # Write category summaries to worksheet
                    for i, row in enumerate(category_summaries.itertuples(), 2):
                        ws_sheet.cell(row=i, column=1, value=row.category)
                        ws_sheet.cell(
                            row=i, column=2, value=row.total_amount
                        ).number_format = '"$"#,##0.00'
                        ws_sheet.cell(row=i, column=3, value=row.count)
                        ws_sheet.cell(
                            row=i, column=4, value=row.average
                        ).number_format = '"$"#,##0.00'

                    # Add a grand total row
                    total_row = len(category_summaries) + 2
                    ws_sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(
                        bold=True
                    )
                    ws_sheet.cell(
                        row=total_row,
                        column=2,
                        value=category_summaries["total_amount"].sum(),
                    ).number_format = '"$"#,##0.00'
                    ws_sheet.cell(
                        row=total_row, column=3, value=category_summaries["count"].sum()
                    ).font = Font(bold=True)

                    # Try to add a pie chart of categories
                    try:
                        chart = PieChart()
                        chart.title = f"{worksheet_name} Categories"

                        # Create references for the chart data - fix the Reference object issue
                        categories_end_row = (
                            len(category_summaries) + 1
                        )  # +1 for header

                        # Use explicit row range references instead of using total_row variable
                        labels = Reference(
                            ws_sheet, min_col=1, min_row=2, max_row=categories_end_row
                        )
                        data_ref = Reference(
                            ws_sheet, min_col=2, min_row=1, max_row=categories_end_row
                        )

                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(labels)
                        chart.height = 10
                        chart.width = 15
                        ws_sheet.add_chart(chart, "F2")
                    except Exception as e:
                        logger.warning(f"Could not create pie chart: {e}")
                        # Log more details for debugging
                        logger.warning(
                            f"Chart error details: {type(e).__name__}, {str(e)}"
                        )

                    logger.info(
                        f"Created summary worksheet '{worksheet_name}' with {len(category_summaries)} categories"
                    )
                else:
                    ws_sheet.cell(row=2, column=1, value="No category data available")

                # Adjust column widths
                for col in ws_sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                    ws_sheet.column_dimensions[column].width = adjusted_width
            else:
                # Create empty worksheet with headers
                ws_sheet = wb.create_sheet(worksheet_name)
                summary_headers = [
                    "Category",
                    "Total Amount",
                    "Count",
                    "Average Amount",
                ]
                for col, header in enumerate(summary_headers, 1):
                    cell = ws_sheet.cell(row=1, column=col, value=header)
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = self.header_alignment
                ws_sheet.cell(
                    row=2,
                    column=1,
                    value=f"No transactions found for {worksheet_name} worksheet",
                )
                logger.info(f"Created empty worksheet '{worksheet_name}'")

        # Create Data Details worksheets for each worksheet type
        # These show the actual transaction data, but categorized by worksheet
        details_folder = wb.create_sheet("_Details")
        details_folder.sheet_state = "hidden"  # Hide the empty details folder

        data["worksheet"] = data["worksheet"].fillna("Unknown").astype(str)
        grouped_by_worksheet = data.groupby("worksheet")

        for worksheet_name in expected_worksheets:
            if worksheet_name in grouped_by_worksheet.groups:
                # Get data for this worksheet
                worksheet_data = grouped_by_worksheet.get_group(worksheet_name)

                # Create a detailed data sheet (hidden by default)
                detail_sheet_name = f"{worksheet_name}_Details"
                ws_details = wb.create_sheet(detail_sheet_name)
                ws_details.sheet_state = "hidden"  # Hidden by default

                # Write headers
                for col, header in enumerate(all_columns, 1):
                    cell = ws_details.cell(row=1, column=col, value=header)
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.alignment = self.header_alignment

                # Write data for this worksheet
                for i, row in enumerate(worksheet_data.itertuples(), 2):
                    for col_idx, value in enumerate(row[1:], 1):  # Skip index
                        ws_details.cell(row=i, column=col_idx, value=value)

                logger.info(
                    f"Created details worksheet '{detail_sheet_name}' with {len(worksheet_data)} transactions"
                )

        # Add data validation for categories and classifications
        # Create hidden validation sheet
        validation_sheet = wb.create_sheet("_Validation")
        validation_sheet.sheet_state = "hidden"

        # Write categories and classifications
        for i, cat in enumerate(categories, 1):
            validation_sheet[f"A{i}"] = cat
        for i, cls in enumerate(classifications, 1):
            validation_sheet[f"B{i}"] = cls

        # Get column letters for validation
        try:
            category_col_letter = get_column_letter(all_columns.index("category") + 1)
            class_col_letter = get_column_letter(
                all_columns.index("classification") + 1
            )

            # Add validation to category column
            cat_validation = DataValidation(
                type="list",
                formula1=f"'\\_Validation'!$A$1:$A${len(categories)}",
                allow_blank=True,
            )
            ws.add_data_validation(cat_validation)
            cat_validation.add(
                f"{category_col_letter}2:{category_col_letter}{len(data)+1}"
            )

            # Add validation to classification column
            class_validation = DataValidation(
                type="list",
                formula1=f"'\\_Validation'!$B$1:$B${len(classifications)}",
                allow_blank=True,
            )
            ws.add_data_validation(class_validation)
            class_validation.add(f"{class_col_letter}2:{class_col_letter}{len(data)+1}")
        except ValueError as e:
            logger.warning(f"Could not add validation - column not found? {e}")

        # Save the Excel workbook
        try:
            wb.save(output_path)
            logger.info(f"Successfully created Excel report at: {output_path}")
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            # Don't proceed to CSV generation if Excel save fails severely
            raise

        # ---- End of Excel Generation ----

        # ---- Start of CSV Generation ----
        # Define the base output directory and create the CSV subdirectory
        base_output_dir = os.path.dirname(output_path)
        csv_output_dir = os.path.join(base_output_dir, "csv_sheets")
        os.makedirs(csv_output_dir, exist_ok=True)
        logger.info(f"CSV output directory: {csv_output_dir}")

        # Make sure expense_type is used to determine Personal items, not just worksheet
        if "expense_type" in data.columns:
            # Force worksheet to be Personal for personal expense types
            personal_mask = data["expense_type"] == "personal"
            if personal_mask.any():
                logger.info(
                    f"Setting worksheet to 'Personal' for {personal_mask.sum()} personal expense items"
                )
                data.loc[personal_mask, "worksheet"] = "Personal"

                # Also ensure business_percentage is set to 0 for personal expenses
                if "business_percentage" in data.columns:
                    data.loc[personal_mask, "business_percentage"] = 0
                    logger.info(
                        f"Set business_percentage to 0 for {personal_mask.sum()} personal expense items"
                    )

        # Additional check for Personal worksheet items to ensure business percentage is 0
        if "worksheet" in data.columns and "business_percentage" in data.columns:
            personal_worksheet_mask = data["worksheet"] == "Personal"
            if personal_worksheet_mask.any():
                # Any item in Personal worksheet should have 0% business
                data.loc[personal_worksheet_mask, "business_percentage"] = 0
                logger.info(
                    f"Ensured business_percentage is 0 for {personal_worksheet_mask.sum()} items in Personal worksheet"
                )

        # Ensure worksheet column contains valid values
        data["worksheet"] = data["worksheet"].fillna("Unknown").astype(str)

        # Group data by the 'worksheet' column
        grouped_data = data.groupby("worksheet")

        # Iterate through each group and save as CSV
        for name, group in grouped_data:
            # Sanitize worksheet name for use in filename
            safe_name = "".join(
                c for c in name if c.isalnum() or c in ("_", "-")
            ).rstrip()
            if not safe_name:  # Handle cases where name becomes empty after sanitizing
                safe_name = "other"

            csv_filename = f"{client_name}_{safe_name}_worksheet.csv"
            csv_filepath = os.path.join(csv_output_dir, csv_filename)

            try:
                # Select only relevant columns if desired, or save all
                group.to_csv(csv_filepath, index=False)
                logger.info(
                    f"Successfully saved CSV for worksheet '{name}' to: {csv_filepath}"
                )
            except Exception as e:
                logger.error(f"Error saving CSV for worksheet '{name}': {e}")

        # Also save a category summary CSV for each worksheet
        for worksheet_name, group in grouped_data:
            if not group.empty and "category" in group.columns:
                # Skip Unknown worksheet
                if worksheet_name == "Unknown":
                    continue

                # Get category summaries
                category_summaries = (
                    group.groupby("category")
                    .agg(
                        total_amount=("amount", "sum"),
                        count=("amount", "count"),
                    )
                    .reset_index()
                )

                # Add average column
                category_summaries["average"] = (
                    category_summaries["total_amount"] / category_summaries["count"]
                )

                # Sort by total amount (descending)
                category_summaries = category_summaries.sort_values(
                    "total_amount", ascending=False
                )

                # Save as CSV
                safe_name = "".join(
                    c for c in worksheet_name if c.isalnum() or c in ("_", "-")
                ).rstrip()
                if not safe_name:
                    safe_name = "other"

                csv_filename = f"{client_name}_{safe_name}_summary.csv"
                csv_filepath = os.path.join(csv_output_dir, csv_filename)

                try:
                    category_summaries.to_csv(csv_filepath, index=False)
                    logger.info(
                        f"Successfully saved summary CSV for worksheet '{worksheet_name}' to: {csv_filepath}"
                    )
                except Exception as e:
                    logger.error(
                        f"Error saving summary CSV for worksheet '{worksheet_name}': {e}"
                    )

        # Additionally, save a CSV of all transactions if useful
        all_csv_filename = f"{client_name}_all_transactions.csv"
        all_csv_filepath = os.path.join(csv_output_dir, all_csv_filename)
        try:
            data.to_csv(all_csv_filepath, index=False)
            logger.info(
                f"Successfully saved CSV for all transactions to: {all_csv_filepath}"
            )
        except Exception as e:
            logger.error(f"Error saving CSV for all transactions: {e}")

        # ---- End of CSV Generation ----

    def save_csv_files(self, data: pd.DataFrame, client_name: str, csv_output_dir: str):
        """Save transactions and summary CSV files for each worksheet."""
        os.makedirs(csv_output_dir, exist_ok=True)
        logger.info(f"CSV output directory: {csv_output_dir}")

        # Ensure data is properly prepared
        # Make sure expense_type is used to determine Personal items, not just worksheet
        if "expense_type" in data.columns:
            # Force worksheet to be Personal for personal expense types
            personal_mask = data["expense_type"] == "personal"
            if personal_mask.any():
                logger.info(
                    f"Setting worksheet to 'Personal' for {personal_mask.sum()} personal expense items"
                )
                data.loc[personal_mask, "worksheet"] = "Personal"

                # Also ensure business_percentage is set to 0 for personal expenses
                if "business_percentage" in data.columns:
                    data.loc[personal_mask, "business_percentage"] = 0
                    logger.info(
                        f"Set business_percentage to 0 for {personal_mask.sum()} personal expense items"
                    )

        # Additional check for Personal worksheet items to ensure business percentage is 0
        if "worksheet" in data.columns and "business_percentage" in data.columns:
            personal_worksheet_mask = data["worksheet"] == "Personal"
            if personal_worksheet_mask.any():
                # Any item in Personal worksheet should have 0% business
                data.loc[personal_worksheet_mask, "business_percentage"] = 0
                logger.info(
                    f"Ensured business_percentage is 0 for {personal_worksheet_mask.sum()} items in Personal worksheet"
                )

        # Ensure worksheet column contains valid values
        data["worksheet"] = data["worksheet"].fillna("Unknown").astype(str)

        # Group data by the 'worksheet' column
        grouped_data = data.groupby("worksheet")

        # Iterate through each group and save as CSV
        for name, group in grouped_data:
            # Sanitize worksheet name for use in filename
            safe_name = "".join(
                c for c in name if c.isalnum() or c in ("_", "-")
            ).rstrip()
            if not safe_name:  # Handle cases where name becomes empty after sanitizing
                safe_name = "other"

            csv_filename = f"{client_name}_{safe_name}_worksheet.csv"
            csv_filepath = os.path.join(csv_output_dir, csv_filename)

            try:
                # Select only relevant columns if desired, or save all
                group.to_csv(csv_filepath, index=False)
                logger.info(
                    f"Successfully saved CSV for worksheet '{name}' to: {csv_filepath}"
                )
            except Exception as e:
                logger.error(f"Error saving CSV for worksheet '{name}': {e}")

        # Also save a category summary CSV for each worksheet
        for worksheet_name, group in grouped_data:
            if not group.empty and "category" in group.columns:
                # Skip Unknown worksheet
                if worksheet_name == "Unknown":
                    continue

                # Get category summaries
                category_summaries = (
                    group.groupby("category")
                    .agg(
                        total_amount=("amount", "sum"),
                        count=("amount", "count"),
                    )
                    .reset_index()
                )

                # Add average column
                category_summaries["average"] = (
                    category_summaries["total_amount"] / category_summaries["count"]
                )

                # Sort by total amount (descending)
                category_summaries = category_summaries.sort_values(
                    "total_amount", ascending=False
                )

                # Save as CSV
                safe_name = "".join(
                    c for c in worksheet_name if c.isalnum() or c in ("_", "-")
                ).rstrip()
                if not safe_name:
                    safe_name = "other"

                csv_filename = f"{client_name}_{safe_name}_summary.csv"
                csv_filepath = os.path.join(csv_output_dir, csv_filename)

                try:
                    category_summaries.to_csv(csv_filepath, index=False)
                    logger.info(
                        f"Successfully saved summary CSV for worksheet '{worksheet_name}' to: {csv_filepath}"
                    )
                except Exception as e:
                    logger.error(
                        f"Error saving summary CSV for worksheet '{worksheet_name}': {e}"
                    )

        # Additionally, save a CSV of all transactions if useful
        all_csv_filename = f"{client_name}_all_transactions.csv"
        all_csv_filepath = os.path.join(csv_output_dir, all_csv_filename)
        try:
            data.to_csv(all_csv_filepath, index=False)
            logger.info(
                f"Successfully saved CSV for all transactions to: {all_csv_filepath}"
            )
        except Exception as e:
            logger.error(f"Error saving CSV for all transactions: {e}")

    def create_linked_excel_report(
        self, data: pd.DataFrame, output_path: str, client_name: str
    ):
        """Create an Excel report with linked formulas instead of raw values.

        This creates a report where summary sheets use formulas that reference the main transactions sheet,
        allowing changes on the transaction sheet to automatically update the summaries.
        """
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        logger.info("Creating Excel report with linked formulas")

        # Prepare data
        data = data.fillna("")

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create main Transactions sheet
        main_sheet = wb.active
        main_sheet.title = "Transactions"

        # Write headers
        all_columns = list(data.columns)
        for col, header in enumerate(all_columns, 1):
            cell = main_sheet.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

        # Write transaction data
        for i, row in enumerate(data.itertuples(), 2):
            for col_idx, value in enumerate(row[1:], 1):  # Skip index
                main_sheet.cell(row=i, column=col_idx, value=value)

        # Get expected worksheets
        expected_worksheets = ["6A", "Vehicle", "HomeOffice", "Personal", "Unknown"]

        # Add the exclusion logic before creating worksheet summaries
        # Apply exclusion logic - move excluded categories from 6A to Personal worksheet
        worksheet_col_idx = all_columns.index("worksheet") + 1
        category_col_idx = (
            all_columns.index("category") + 1 if "category" in all_columns else None
        )

        if category_col_idx:
            for row in range(2, len(data) + 2):  # +2 because row 1 is header
                category = main_sheet.cell(row=row, column=category_col_idx).value
                worksheet = main_sheet.cell(row=row, column=worksheet_col_idx).value

                if worksheet == "6A" and category in EXCLUDED_FROM_6A:
                    # Move to Personal worksheet
                    main_sheet.cell(row=row, column=worksheet_col_idx, value="Personal")
                    logger.info(f"Excluded '{category}' from 6A worksheet (row: {row})")

        # Create worksheet summaries with linked formulas
        amount_col_idx = (
            all_columns.index("amount") + 1 if "amount" in all_columns else None
        )

        if amount_col_idx is None:
            logger.error("Cannot create linked report - 'amount' column is missing")
            return

        # For each worksheet type
        for worksheet_name in expected_worksheets:
            if worksheet_name == "Unknown":
                continue

            # Create worksheet summary sheet
            ws_sheet = wb.create_sheet(worksheet_name)

            # Set up headers
            summary_headers = [
                "Category",
                "Total Amount",
                "Count",
                "Average Amount",
            ]
            for col, header in enumerate(summary_headers, 1):
                cell = ws_sheet.cell(row=1, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.header_alignment

            # Get all unique categories in this worksheet
            categories = set()
            for row in range(2, len(data) + 2):
                if (
                    main_sheet.cell(row=row, column=worksheet_col_idx).value
                    == worksheet_name
                ):
                    if category_col_idx:
                        cat = main_sheet.cell(row=row, column=category_col_idx).value
                        if cat:
                            categories.add(cat)

            # Sort categories
            categories = sorted(list(categories))

            # Add formula-based summaries for each category
            for i, category in enumerate(categories, 2):
                # Category name
                ws_sheet.cell(row=i, column=1, value=category)

                # Total amount (SUMIFS formula)
                total_formula = (
                    f"=SUMIFS(Transactions!{get_column_letter(amount_col_idx)}:"
                    f"{get_column_letter(amount_col_idx)},"
                    f"Transactions!{get_column_letter(worksheet_col_idx)}:"
                    f'{get_column_letter(worksheet_col_idx)},"{worksheet_name}",'
                    f"Transactions!{get_column_letter(category_col_idx)}:"
                    f'{get_column_letter(category_col_idx)},"{category}")'
                )
                ws_sheet.cell(row=i, column=2, value=total_formula).number_format = (
                    '"$"#,##0.00'
                )

                # Count (COUNTIFS formula)
                count_formula = (
                    f"=COUNTIFS(Transactions!{get_column_letter(worksheet_col_idx)}:"
                    f'{get_column_letter(worksheet_col_idx)},"{worksheet_name}",'
                    f"Transactions!{get_column_letter(category_col_idx)}:"
                    f'{get_column_letter(category_col_idx)},"{category}")'
                )
                ws_sheet.cell(row=i, column=3, value=count_formula)

                # Average amount (formula referencing other cells)
                avg_formula = f"=IF({get_column_letter(3)}{i}>0,{get_column_letter(2)}{i}/{get_column_letter(3)}{i},0)"
                ws_sheet.cell(row=i, column=4, value=avg_formula).number_format = (
                    '"$"#,##0.00'
                )

            # Add totals row with formulas
            total_row = len(categories) + 2
            ws_sheet.cell(row=total_row, column=1, value="TOTAL").font = (
                openpyxl.styles.Font(bold=True)
            )

            # Sum of all amounts
            ws_sheet.cell(
                row=total_row,
                column=2,
                value=f"=SUM({get_column_letter(2)}2:{get_column_letter(2)}{total_row-1})",
            ).number_format = '"$"#,##0.00'

            # Sum of all counts
            ws_sheet.cell(
                row=total_row,
                column=3,
                value=f"=SUM({get_column_letter(3)}2:{get_column_letter(3)}{total_row-1})",
            ).font = openpyxl.styles.Font(bold=True)

            # Adjust column widths
            for col in ws_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                ws_sheet.column_dimensions[column].width = adjusted_width

            # Add autofilter to header row
            main_sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(all_columns))}{len(data)+1}"
            )

            logger.info(
                f"Created linked worksheet summary for '{worksheet_name}' with {len(categories)} categories"
            )

        # Create a Totals Summary sheet
        totals_sheet = wb.create_sheet("Summary")
        totals_sheet.cell(row=1, column=1, value="Worksheet").fill = self.header_fill
        totals_sheet.cell(row=1, column=2, value="Total Amount").fill = self.header_fill
        totals_sheet.cell(row=1, column=3, value="Transaction Count").fill = (
            self.header_fill
        )

        # Add formulas that reference each worksheet's total
        for i, worksheet_name in enumerate(expected_worksheets, 2):
            if worksheet_name == "Unknown":
                continue

            totals_sheet.cell(row=i, column=1, value=worksheet_name)
            # Reference the total cell in each worksheet
            if worksheet_name in wb.sheetnames:
                total_cell_ref = (
                    f"'{worksheet_name}'!B{len(categories)+2}"  # Reference total cell
                )
                count_cell_ref = (
                    f"'{worksheet_name}'!C{len(categories)+2}"  # Reference count cell
                )

                totals_sheet.cell(
                    row=i, column=2, value=f"={total_cell_ref}"
                ).number_format = '"$"#,##0.00'
                totals_sheet.cell(row=i, column=3, value=f"={count_cell_ref}")
            else:
                totals_sheet.cell(row=i, column=2, value=0).number_format = (
                    '"$"#,##0.00'
                )
                totals_sheet.cell(row=i, column=3, value=0)

        # Add grand total row
        grand_total_row = (
            len(expected_worksheets) - 1 + 2
        )  # -1 for Unknown, +2 for header and adjustment
        totals_sheet.cell(row=grand_total_row, column=1, value="GRAND TOTAL").font = (
            openpyxl.styles.Font(bold=True)
        )
        totals_sheet.cell(
            row=grand_total_row, column=2, value=f"=SUM(B2:B{grand_total_row-1})"
        ).number_format = '"$"#,##0.00'
        totals_sheet.cell(
            row=grand_total_row, column=3, value=f"=SUM(C2:C{grand_total_row-1})"
        ).font = openpyxl.styles.Font(bold=True)

        # Adjust column widths
        for col in totals_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50
            totals_sheet.column_dimensions[column].width = adjusted_width

        # Save the workbook
        try:
            wb.save(output_path)
            logger.info(f"Successfully created linked Excel report at: {output_path}")
        except Exception as e:
            logger.error(f"Error saving linked Excel file: {e}")
            raise
